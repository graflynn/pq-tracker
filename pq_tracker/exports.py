from __future__ import annotations

import json
import logging
import re
import sqlite3
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from . import config as cfg
from . import db

log = logging.getLogger(__name__)

XLSX_COLUMNS = [
    ("pq_ref", "PQ Ref", 14),
    ("date_asked", "Date Asked", 12),
    ("date_answered", "Date Answered", 12),
    ("td_name", "TD", 24),
    ("td_party", "Party", 14),
    ("td_constituency", "Constituency", 22),
    ("constituent", "Constituent (manual)", 24),
    ("tags_csv", "Tags (manual)", 24),
    ("notes", "Notes (manual)", 32),
    ("department", "Dept", 14),
    ("minister_name", "Minister", 22),
    ("matched_topics", "Matched Topics", 28),
    ("answer_status", "Status", 10),
    ("question_text", "Question", 80),
    ("answer_text", "Answer", 80),
    ("oireachtas_permalink", "Permalink", 50),
    ("hse_pdf_url", "HSE PDF (manual)", 30),
]


def write_xlsx(rows: list[sqlite3.Row], dest,
               tags_map: dict[str, list[str]] | None = None,
               answers_map: dict[str, str] | None = None) -> None:
    """Write the xlsx to ``dest``, which can be a Path or a writable file-like
    object (e.g. an in-memory BytesIO for streaming as an HTTP download).

    ``answers_map`` (pq_ref → answer_text) is required for the Answer column
    since the text lives on `answers`, not `questions`. Pass
    ``db.answers_by_pqref(conn)``.
    """
    tags_map = tags_map or {}
    answers_map = answers_map or {}
    if isinstance(dest, Path):
        dest.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "PQs"
    header_font = Font(bold=True)
    for col_idx, (_, label, width) in enumerate(XLSX_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (key, _, _) in enumerate(XLSX_COLUMNS, start=1):
            if key == "tags_csv":
                value = ", ".join(tags_map.get(row["pq_ref"], []))
            elif key == "answer_text":
                value = answers_map.get(row["pq_ref"])
            else:
                value = row[key] if key in row.keys() else None
                if key == "matched_topics" and value:
                    try:
                        value = ", ".join(json.loads(value))
                    except (TypeError, json.JSONDecodeError):
                        pass
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(dest)
    if isinstance(dest, Path):
        log.info("wrote %s (%d rows)", dest, len(rows))


_FILENAME_SAFE = re.compile(r"[^A-Za-z0-9._-]+")


def _safe_pq_ref_for_filename(pq_ref: str) -> str:
    # "16475/26" -> "16475-26"
    return _FILENAME_SAFE.sub("-", pq_ref)


def write_question_markdown(row: sqlite3.Row, out_dir: Path,
                            tags: list[str] | None = None,
                            answer_text: str | None = None) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    safe = _safe_pq_ref_for_filename(row["pq_ref"])
    date_asked = row["date_asked"] or ""
    path = out_dir / f"PQ-{safe}-{date_asked}.md"
    try:
        topics = ", ".join(json.loads(row["matched_topics"] or "[]"))
    except json.JSONDecodeError:
        topics = row["matched_topics"] or ""
    lines: list[str] = []
    lines.append(f"# PQ {row['pq_ref']}")
    lines.append("")
    lines.append(f"- **Date asked:** {row['date_asked']}")
    lines.append(f"- **Date answered:** {row['date_answered'] or '_pending_'}")
    lines.append(f"- **TD:** {row['td_name']}"
                 + (f" ({row['td_party']})" if row['td_party'] else "")
                 + (f" — {row['td_constituency']}" if row['td_constituency'] else ""))
    lines.append(f"- **Department:** {row['department']}")
    if row["minister_name"]:
        lines.append(f"- **Minister:** {row['minister_name']}")
    lines.append(f"- **Status:** {row['answer_status']}")
    lines.append(f"- **Matched topics:** {topics}")
    if tags:
        lines.append(f"- **Tags:** {', '.join(tags)}")
    if row["constituent"]:
        lines.append(f"- **Constituent:** {row['constituent']}")
    lines.append(f"- **Permalink:** {row['oireachtas_permalink']}")
    if row["hse_pdf_url"]:
        lines.append(f"- **HSE PDF (manual):** {row['hse_pdf_url']}")
    if row["notes"]:
        lines.append("")
        lines.append("## Notes")
        lines.append("")
        lines.append(row["notes"])
    lines.append("")
    lines.append("## Question")
    lines.append("")
    lines.append(row["question_text"] or "")
    lines.append("")
    lines.append("## Answer")
    lines.append("")
    if answer_text:
        lines.append(answer_text)
    else:
        lines.append("_Answer not yet published._")
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def write_summary(path: Path, *, run_started_iso: str,
                  new_questions: int, newly_answered: int, errors: int,
                  topics: list[str], date_start: date, date_end: date,
                  rows_in_run: list[sqlite3.Row], total_rows: int,
                  pending_total: int) -> None:
    td_counter: Counter = Counter()
    topic_counter: Counter = Counter()
    for r in rows_in_run:
        td_counter[r["td_name"]] += 1
        try:
            for t in json.loads(r["matched_topics"] or "[]"):
                topic_counter[t] += 1
        except json.JSONDecodeError:
            pass
    lines = [
        "# pq-tracker — latest run",
        "",
        f"- **Run started (UTC):** {run_started_iso}",
        f"- **Window:** {date_start} → {date_end}",
        f"- **Keywords:** {', '.join(topics)}",
        "",
        "## Counts",
        "",
        f"- New questions ingested this run: **{new_questions}**",
        f"- Newly answered this run: **{newly_answered}**",
        f"- Errors during run: **{errors}**",
        f"- Total rows in DB: **{total_rows}**",
        f"- Pending (no answer yet): **{pending_total}**",
        "",
        "## Top TDs in this run",
        "",
    ]
    if td_counter:
        for name, n in td_counter.most_common(10):
            lines.append(f"- {name}: {n}")
    else:
        lines.append("_(none)_")
    lines += ["", "## Top topics in this run", ""]
    if topic_counter:
        for t, n in topic_counter.most_common(10):
            lines.append(f"- {t}: {n}")
    else:
        lines.append("_(none)_")
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")
    log.info("wrote %s", path)


def write_all(conn: sqlite3.Connection, run_started_iso: str,
              new_questions: int, newly_answered: int, errors: int,
              topics: list[str], date_start: date, date_end: date) -> None:
    all_rows = db.all_questions(conn)
    rows_in_run = db.all_matches_in_run(conn, run_started_iso)
    pending_total = sum(1 for r in all_rows if r["answer_status"] == "pending")
    tags_map = db.tags_by_pqref(conn)
    answers_map = db.answers_by_pqref(conn)

    # The xlsx is now an on-demand export ("Export to Excel" in the UI), so the
    # ingest run no longer regenerates a stale full-corpus snapshot here.
    for row in all_rows:
        write_question_markdown(
            row, cfg.QUESTIONS_DIR,
            tags=tags_map.get(row["pq_ref"]),
            answer_text=answers_map.get(row["pq_ref"]),
        )
    write_summary(
        cfg.SUMMARY_PATH,
        run_started_iso=run_started_iso,
        new_questions=new_questions,
        newly_answered=newly_answered,
        errors=errors,
        topics=topics,
        date_start=date_start,
        date_end=date_end,
        rows_in_run=rows_in_run,
        total_rows=len(all_rows),
        pending_total=pending_total,
    )
